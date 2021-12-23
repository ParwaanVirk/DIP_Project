import openpyxl
from backend.models import Transformer
import dateutil.parser
def createTransformer(TimeStamp, Transformer_ID, Locality, Current_Input, 
Voltage_Input, Oil_Temprature, Winding_Temprature, Oil_TankLevel, Moisture_Level, 
Tapping_Ratio):

    TransformerEntry = Transformer.objects.create(
        TimeStamp = TimeStamp, 
        Transformer_ID = Transformer_ID, 
        Locality = Locality, 
        Current_Input = Current_Input, 
        Voltage_Input = Voltage_Input, 
        Oil_Temprature = Oil_Temprature, 
        Winding_Temprature = Winding_Temprature, 
        Oil_TankLevel = Oil_TankLevel, 
        Moisture_Level = Moisture_Level,
        Tapping_Ratio = Tapping_Ratio, 
        Overall_Health = 0

    )


def createTransFromXLSheet(xl_book, sheet_name):
    my_bk = openpyxl.load_workbook(xl_book)
    my_sheet = my_bk[sheet_name]
    row = 2
    while(my_sheet.cell(row=row, column=1).value != None):
        k = str(my_sheet.cell(row=row, column=1).value)
        l = dateutil.parser.parse(k)
        timestampgen = l
        transformerid = "TF-03"
        locality = "02"
        currentinput = my_sheet.cell(row=row, column=2).value
        voltageinput = my_sheet.cell(row=row, column=3).value
        oiltemp = my_sheet.cell(row=row, column=4).value
        windtemp = my_sheet.cell(row=row, column=5).value
        oillevel = my_sheet.cell(row=row, column=6).value
        moisturelevel = my_sheet.cell(row=row, column=7).value
        tappingratio = my_sheet.cell(row=row, column=8).value
        createTransformer(timestampgen, transformerid, locality, currentinput, voltageinput, 
        oiltemp, windtemp, oillevel, moisturelevel, tappingratio)
        print("NEW Entry")
        row +=1
    print("Entry for 1 transformer done")


createTransFromXLSheet("./Sheets/DIP4-Health Index Estimation Training Data.xlsx", "TF-03")